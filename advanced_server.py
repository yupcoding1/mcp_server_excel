import os
from dotenv import load_dotenv
from openpyxl.utils.exceptions import InvalidFileException
from excel_fucntion import *
from contextlib import asynccontextmanager
from collections.abc import AsyncGenerator
from mcp.server import Server
import mcp.server.stdio
import mcp.types as types
from mcp.server.lowlevel import NotificationOptions
from mcp.server.models import InitializationOptions

load_dotenv()
EXCEL_FILES_DIR = os.getenv("EXCEL_FILES_DIR", "./excel_files")

@asynccontextmanager
async def server_lifespan(server: Server) -> AsyncGenerator[dict, None]:
    os.makedirs(EXCEL_FILES_DIR, exist_ok=True)
    yield {"excel_dir": EXCEL_FILES_DIR}

server = Server("excel-advanced-server", lifespan=server_lifespan)

@server.list_resources()
async def handle_list_resources(name: str, arguments: dict | None) -> list[types.Resource]:
    files = [f for f in os.listdir(EXCEL_FILES_DIR) if f.endswith(".xlsx") or f.endswith(".xlsm")]
    return [
        types.Resource(
            name=f"excel-file://{file}",
            description=f"Excel file: {file}",
            arguments=[]
        ) for file in files
    ]

@server.read_resource()
async def handle_get_resource(name: str, arguments: dict | None) -> types.ReadResourceResult:
    if name.startswith("excel-file://"):
        filename = name.replace("excel-file://", "")
        path = os.path.join(EXCEL_FILES_DIR, filename)
        try:
            sheets = list_sheets(path)
            return types.ReadResourceResult(
                description=f"Sheets in {filename}",
                content=types.TextContent(type="text", text="\n".join(sheets)),
                mime_type="text/plain"
            )
        except (FileNotFoundError, InvalidFileException) as e:
            return types.ReadResourceResult(
                description=f"Error: {str(e)}",
                content=types.TextContent(type="text", text=str(e)),
                mime_type="text/plain"
            )
    raise ValueError(f"Unknown resource: {name}")

@server.list_tools()
async def handle_list_tools() -> list[types.Tool]:
    return [
        types.Tool(
            name="create_excel_file",
            description="Create a new Excel file.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "File name"},
                    "sheet_name": {"type": "string", "description": "Sheet name"}
                },
                "required": ["filename"]
            }
        ),
        types.Tool(
            name="add_sheet",
            description="Add a new sheet to an Excel file.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "File name"},
                    "sheet_name": {"type": "string", "description": "Sheet name"}
                },
                "required": ["filename", "sheet_name"]
            }
        ),
        types.Tool(
            name="rename_sheet",
            description="Rename a sheet in an Excel file.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "File name"},
                    "old_name": {"type": "string", "description": "Old sheet name"},
                    "new_name": {"type": "string", "description": "New sheet name"}
                },
                "required": ["filename", "old_name", "new_name"]
            }
        ),
        types.Tool(
            name="delete_sheet",
            description="Delete a sheet from an Excel file.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "File name"},
                    "sheet_name": {"type": "string", "description": "Sheet name"}
                },
                "required": ["filename", "sheet_name"]
            }
        ),
        types.Tool(
            name="write_cell",
            description="Write a value to a cell.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "File name"},
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "cell": {"type": "string", "description": "Cell address"},
                    "value": {"type": "string", "description": "Value to write"}
                },
                "required": ["filename", "sheet", "cell", "value"]
            }
        ),
        types.Tool(
            name="read_cell",
            description="Read a value from a cell.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "File name"},
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "cell": {"type": "string", "description": "Cell address"}
                },
                "required": ["filename", "sheet", "cell"]
            }
        ),
        types.Tool(
            name="merge_cells",
            description="Merge a range of cells.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "File name"},
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "cell_range": {"type": "string", "description": "Cell range (e.g. A1:B2)"}
                },
                "required": ["filename", "sheet", "cell_range"]
            }
        ),
        types.Tool(
            name="unmerge_cells",
            description="Unmerge a range of cells.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "File name"},
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "cell_range": {"type": "string", "description": "Cell range (e.g. A1:B2)"}
                },
                "required": ["filename", "sheet", "cell_range"]
            }
        ),
        types.Tool(
            name="write_row",
            description="Write a row of data starting at a cell.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "File name"},
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "start_cell": {"type": "string", "description": "Start cell address"},
                    "data": {"type": "array", "description": "List of values", "items": {"type": "string"}}
                },
                "required": ["filename", "sheet", "start_cell", "data"]
            }
        ),
        types.Tool(
            name="write_column",
            description="Write a column of data starting at a cell.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "File name"},
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "start_cell": {"type": "string", "description": "Start cell address"},
                    "data": {"type": "array", "description": "List of values", "items": {"type": "string"}}
                },
                "required": ["filename", "sheet", "start_cell", "data"]
            }
        ),
        types.Tool(
            name="set_border",
            description="Set border for a range of cells.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "File name"},
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "cell_range": {"type": "string", "description": "Cell range (e.g. A1:B2)"}
                },
                "required": ["filename", "sheet", "cell_range"]
            }
        ),
        types.Tool(
            name="auto_fit_columns",
            description="Auto-fit columns in a sheet.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "File name"},
                    "sheet": {"type": "string", "description": "Sheet name"}
                },
                "required": ["filename", "sheet"]
            }
        ),
        types.Tool(
            name="get_used_range",
            description="Get the used range of a sheet.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "File name"},
                    "sheet": {"type": "string", "description": "Sheet name"}
                },
                "required": ["filename", "sheet"]
            }
        ),
        types.Tool(
            name="read_range",
            description="Read a range of cells.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "File name"},
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "cell_range": {"type": "string", "description": "Cell range (e.g. A1:B2)"}
                },
                "required": ["filename", "sheet", "cell_range"]
            }
        ),
        types.Tool(
            name="write_formula",
            description="Write a formula to a cell.",
            inputSchema={
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "File name"},
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "cell": {"type": "string", "description": "Cell address"},
                    "formula": {"type": "string", "description": "Formula string (without =)"}
                },
                "required": ["filename", "sheet", "cell", "formula"]
            }
        ),
        types.Tool(
            name="save_as_new_file",
            description="Save the Excel file as a new file.",
            inputSchema={
                "type": "object",
                "properties": {
                    "old_filename": {"type": "string", "description": "Old file name"},
                    "new_filename": {"type": "string", "description": "New file name"}
                },
                "required": ["old_filename", "new_filename"]
            }
        )
    ]

@server.call_tool()
async def handle_call_tool(name: str, arguments: dict) -> list[types.TextContent]:
    path = os.path.join(EXCEL_FILES_DIR, arguments.get("filename", ""))
    if name == "create_excel_file":
        result = create_excel_file(path, arguments.get("sheet_name", "Sheet1"))
    elif name == "add_sheet":
        result = add_sheet(path, arguments["sheet_name"])
    elif name == "rename_sheet":
        result = rename_sheet(path, arguments["old_name"], arguments["new_name"])
    elif name == "delete_sheet":
        result = delete_sheet(path, arguments["sheet_name"])
    elif name == "write_cell":
        result = write_cell(path, arguments["sheet"], arguments["cell"], arguments["value"])
    elif name == "read_cell":
        result = str(read_cell(path, arguments["sheet"], arguments["cell"]))
    elif name == "merge_cells":
        result = merge_cells(path, arguments["sheet"], arguments["cell_range"])
    elif name == "unmerge_cells":
        result = unmerge_cells(path, arguments["sheet"], arguments["cell_range"])
    elif name == "write_row":
        result = write_row(path, arguments["sheet"], arguments["start_cell"], arguments["data"])
    elif name == "write_column":
        result = write_column(path, arguments["sheet"], arguments["start_cell"], arguments["data"])
    elif name == "set_border":
        result = set_border(path, arguments["sheet"], arguments["cell_range"])
    elif name == "auto_fit_columns":
        result = auto_fit_columns(path, arguments["sheet"])
    elif name == "get_used_range":
        result = str(get_used_range(path, arguments["sheet"]))
    elif name == "read_range":
        result = str(read_range(path, arguments["sheet"], arguments["cell_range"]))
    elif name == "write_formula":
        result = write_formula(path, arguments["sheet"], arguments["cell"], arguments["formula"])
    elif name == "save_as_new_file":
        old_path = os.path.join(EXCEL_FILES_DIR, arguments["old_filename"])
        new_path = os.path.join(EXCEL_FILES_DIR, arguments["new_filename"])
        result = save_as_new_file(old_path, new_path)
    else:
        result = f"Unknown tool: {name}"
    return [types.TextContent(type="text", text=str(result))]

async def run():
    async with mcp.server.stdio.stdio_server() as (read_stream, write_stream):
        await server.run(
            read_stream,
            write_stream,
            InitializationOptions(
                server_name="excel-advanced-server",
                server_version="0.1.0",
                capabilities=server.get_capabilities(
                    notification_options=NotificationOptions(),
                    experimental_capabilities={},
                ),
            ),
        )

if __name__ == "__main__":
    import asyncio
    asyncio.run(run())
