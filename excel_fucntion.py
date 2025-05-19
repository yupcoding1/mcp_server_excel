import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import List, Dict, Any, Optional

# ---------- BASIC UTILITIES ----------

def create_excel_file(filename: str, sheet_name: str = "Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    wb.save(filename)
    return f"Created {filename} with sheet '{sheet_name}'"

def load_excel_file(filename: str):
    if not os.path.exists(filename):
        raise FileNotFoundError(f"{filename} does not exist.")
    return load_workbook(filename)

# ---------- SHEET MANAGEMENT ----------

def add_sheet(filename: str, sheet_name: str):
    wb = load_excel_file(filename)
    wb.create_sheet(title=sheet_name)
    wb.save(filename)
    return f"Added sheet '{sheet_name}'"

def rename_sheet(filename: str, old_name: str, new_name: str):
    wb = load_excel_file(filename)
    wb[old_name].title = new_name
    wb.save(filename)
    return f"Renamed sheet from '{old_name}' to '{new_name}'"

def delete_sheet(filename: str, sheet_name: str):
    wb = load_excel_file(filename)
    del wb[sheet_name]
    wb.save(filename)
    return f"Deleted sheet '{sheet_name}'"

# ---------- CELL OPERATIONS ----------

def write_cell(filename: str, sheet: str, cell: str, value: Any, 
               bold=False, italic=False, font_color="000000", bg_color=None, align="left"):
    wb = load_excel_file(filename)
    ws = wb[sheet]
    cell_obj = ws[cell]
    cell_obj.value = value
    cell_obj.font = Font(bold=bold, italic=italic, color=font_color)
    cell_obj.alignment = Alignment(horizontal=align)
    if bg_color:
        cell_obj.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    wb.save(filename)
    return f"Wrote value '{value}' to {cell} in '{sheet}'"

def read_cell(filename: str, sheet: str, cell: str):
    wb = load_excel_file(filename)
    value = wb[sheet][cell].value
    return value

def merge_cells(filename: str, sheet: str, cell_range: str):
    wb = load_excel_file(filename)
    ws = wb[sheet]
    ws.merge_cells(cell_range)
    wb.save(filename)
    return f"Merged cells {cell_range}"

def unmerge_cells(filename: str, sheet: str, cell_range: str):
    wb = load_excel_file(filename)
    ws = wb[sheet]
    ws.unmerge_cells(cell_range)
    wb.save(filename)
    return f"Unmerged cells {cell_range}"

# ---------- ROW/COLUMN BULK OPERATIONS ----------

def write_row(filename: str, sheet: str, start_cell: str, data: List[Any]):
    wb = load_excel_file(filename)
    ws = wb[sheet]
    row = ws[start_cell].row
    col = ws[start_cell].column
    for i, val in enumerate(data):
        ws.cell(row=row, column=col + i, value=val)
    wb.save(filename)
    return f"Wrote row starting at {start_cell}"

def write_column(filename: str, sheet: str, start_cell: str, data: List[Any]):
    wb = load_excel_file(filename)
    ws = wb[sheet]
    row = ws[start_cell].row
    col = ws[start_cell].column
    for i, val in enumerate(data):
        ws.cell(row=row + i, column=col, value=val)
    wb.save(filename)
    return f"Wrote column starting at {start_cell}"

# ---------- FORMATTING UTILITIES ----------

def set_border(filename: str, sheet: str, cell_range: str):
    wb = load_excel_file(filename)
    ws = wb[sheet]
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border
    wb.save(filename)
    return f"Applied border to {cell_range}"

def auto_fit_columns(filename: str, sheet: str):
    wb = load_excel_file(filename)
    ws = wb[sheet]
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(filename)
    return f"Auto-fitted columns in '{sheet}'"

# ---------- SHEET INSPECTION ----------

def list_sheets(filename: str) -> List[str]:
    wb = load_excel_file(filename)
    return wb.sheetnames

def get_used_range(filename: str, sheet: str) -> Dict[str, str]:
    wb = load_excel_file(filename)
    ws = wb[sheet]
    return {
        "min_row": ws.min_row,
        "max_row": ws.max_row,
        "min_col": ws.min_column,
        "max_col": ws.max_column
    }

def read_range(filename: str, sheet: str, cell_range: str) -> List[List[Any]]:
    wb = load_excel_file(filename)
    ws = wb[sheet]
    data = [[cell.value for cell in row] for row in ws[cell_range]]
    return data

# ---------- FORMULA SUPPORT ----------

def write_formula(filename: str, sheet: str, cell: str, formula: str):
    wb = load_excel_file(filename)
    ws = wb[sheet]
    ws[cell] = f"={formula}"
    wb.save(filename)
    return f"Wrote formula '{formula}' in {cell}"

# ---------- SAVE/EXPORT ----------

def save_as_new_file(old_filename: str, new_filename: str):
    wb = load_excel_file(old_filename)
    wb.save(new_filename)
    return f"Saved copy as {new_filename}"
